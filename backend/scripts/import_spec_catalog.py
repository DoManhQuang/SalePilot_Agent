"""Import the multi-category product spreadsheet into MongoDB.

Reads every sheet in ``Spec_cate_gia.xlsx``, normalizes each row through the
category registry, upserts the documents into the Mongo ``products`` collection
(with useful indexes), and writes a JSON snapshot used as an offline fallback.

Usage (from backend/):
    python -m scripts.import_spec_catalog --excel ../Spec_cate_gia.xlsx
    python -m scripts.import_spec_catalog --snapshot-only   # no Mongo needed
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import openpyxl

from app.catalog import repository
from app.catalog.categories import BY_SHEET, Category, normalize_product
from app.config import get_settings

DEFAULT_EXCEL = Path(__file__).resolve().parents[2] / "Spec_cate_gia.xlsx"


def _rows(ws) -> list[dict[str, Any]]:
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []
    for offset, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or all(v in (None, "") for v in values):
            continue
        row = {header[i]: values[i] for i in range(min(len(header), len(values)))}
        row["__row__"] = offset
        rows.append(row)
    return rows


def build_catalog(excel: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, int]]]:
    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    products: list[dict[str, Any]] = []
    stats: dict[str, dict[str, int]] = {}
    seen_skus: set[str] = set()

    for sheet in wb.sheetnames:
        cat: Category | None = BY_SHEET.get(sheet)
        if cat is None:
            continue
        source = f"spec_cate_gia.xlsx:{sheet}"
        total = priced = 0
        for row in _rows(wb[sheet]):
            sku = str(row.get("sku") or "").strip()
            if not sku or sku in seen_skus:
                continue
            doc = normalize_product(cat, row, row["__row__"], source)
            seen_skus.add(sku)
            products.append(doc)
            total += 1
            priced += int(doc["has_current_price"])
        stats[cat.slug] = {"code": cat.code, "total": total, "priced": priced}
    wb.close()
    return products, stats


def write_mongo(products: list[dict[str, Any]]) -> int:
    from pymongo import ASCENDING, UpdateOne

    settings = get_settings()
    client = repository.mongo_client(timeout_ms=4000)
    coll = client[settings.mongodb_db][settings.mongodb_products_collection]
    coll.create_index([("sku", ASCENDING)], unique=True)
    coll.create_index([("category_code", ASCENDING)])
    coll.create_index([("brand", ASCENDING)])
    coll.create_index([("price_vnd", ASCENDING)])
    coll.create_index([("category_code", ASCENDING), ("has_current_price", ASCENDING)])

    ops = [UpdateOne({"sku": doc["sku"]}, {"$set": doc}, upsert=True) for doc in products]
    written = 0
    for start in range(0, len(ops), 1000):
        result = coll.bulk_write(ops[start : start + 1000], ordered=False)
        written += (result.upserted_count or 0) + (result.modified_count or 0)
    # Drop any SKUs no longer present in the source.
    current = {doc["sku"] for doc in products}
    coll.delete_many({"sku": {"$nin": list(current)}})
    count = coll.count_documents({})
    client.close()
    return count


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--snapshot-only", action="store_true", help="Skip MongoDB write")
    args = parser.parse_args()

    if not args.excel.exists():
        raise SystemExit(f"Excel not found: {args.excel}")

    products, stats = build_catalog(args.excel)
    if not products:
        raise SystemExit("No products parsed — check sheet names against the registry.")

    snapshot = repository.save_snapshot(products)
    print(f"Snapshot: {snapshot} ({len(products)} products)")

    if not args.snapshot_only:
        try:
            count = write_mongo(products)
            print(f"MongoDB: upserted into '{get_settings().mongodb_products_collection}', now {count} docs")
        except Exception as exc:  # pragma: no cover - operational path
            print(f"MongoDB write skipped ({exc}). Snapshot is still available for offline use.")

    print("\nPer-category:")
    for slug, s in sorted(stats.items(), key=lambda kv: -kv[1]["total"]):
        print(f"  {slug:18s} code={s['code']:<4} total={s['total']:<5} priced={s['priced']}")
    print(f"\nTOTAL {len(products)} products across {len(stats)} categories")


if __name__ == "__main__":
    main()
